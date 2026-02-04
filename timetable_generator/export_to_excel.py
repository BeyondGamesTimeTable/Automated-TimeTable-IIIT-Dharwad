"""Export timetables to Excel with color coding"""
import pandas as pd
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class TimetableExcelExporter:
    def __init__(self, input_dir='timetable_outputs'):
        self.input_dir = input_dir
        self.course_colors = {}
        self.color_index = 0
        
    def _get_course_color(self, course_code):
        """Generate a consistent color for a course code"""
        if course_code in self.course_colors:
            return self.course_colors[course_code]
        
        # Predefined pleasant colors for courses
        color_palette = [
            'FFE5CC',  # Light Orange
            'FFD9CC',  # Peach
            'FFCCCC',  # Light Pink
            'FFCCF2',  # Light Magenta
            'E5CCFF',  # Light Purple
            'CCD9FF',  # Light Blue
            'CCE5FF',  # Sky Blue
            'CCF2FF',  # Light Cyan
            'CCFFFF',  # Very Light Cyan
            'CCFFE5',  # Mint
            'CCFFD9',  # Light Green
            'D9FFCC',  # Lime
            'E5FFCC',  # Yellow-Green
            'F2FFCC',  # Light Yellow
            'FFFFCC',  # Pale Yellow
            'FFF2CC',  # Cream
            'FFE5B3',  # Light Gold
            'FFD9B3',  # Light Tan
            'FFCCB3',  # Light Salmon
            'FFB3CC',  # Pink
        ]
        
        # Use hash to consistently assign colors
        hash_value = hash(course_code)
        color = color_palette[abs(hash_value) % len(color_palette)]
        
        self.course_colors[course_code] = color
        return color
    
    def _extract_course_code(self, cell_value):
        """Extract course code from cell value"""
        if not isinstance(cell_value, str) or cell_value in ['Free', 'LUNCH BREAK', '']:
            return None
        
        # Handle different formats:
        # "CS162 | C004"
        # "CS162-T | C104"
        # "CS162 (Common) | C004"
        # "CS162-Lab | L106 & L107"
        parts = str(cell_value).split('|')[0].strip()
        
        # Remove suffixes
        parts = parts.replace(' (Common)', '').replace('-T', '').replace('-Lab', '').strip()
        
        # Extract course code (alphanumeric at start)
        import re
        match = re.match(r'^([A-Z]+\d+)', parts)
        if match:
            return match.group(1)
        
        return None
    
    def export_to_excel(self, output_file='All_Timetables.xlsx'):
        """Export all timetables to a single Excel file with multiple sheets"""
        # Find all CSV files
        csv_files = []
        for file in os.listdir(self.input_dir):
            if file.endswith('.csv') and 'Timetable' in file and not file.endswith('_Electives.json'):
                csv_files.append(os.path.join(self.input_dir, file))
        
        if not csv_files:
            print("No timetable CSV files found!")
            return False
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        print(f"\nExporting {len(csv_files)} timetables to Excel...")
        
        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for csv_file in sorted(csv_files):
            filename = Path(csv_file).stem
            sheet_name = filename.replace('_Timetable', '').replace('_', ' ')
            
            # Excel sheet names have max 31 characters
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
            
            print(f"Processing: {sheet_name}")
            
            # Read CSV
            df = pd.read_csv(csv_file, index_col=0)
            
            # Create sheet
            ws = wb.create_sheet(title=sheet_name)
            
            # Write data
            for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Style header row
                    if r_idx == 1:
                        cell.font = Font(bold=True, color='FFFFFF', size=11)
                        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    # Style index column (days)
                    elif c_idx == 1 and r_idx > 1:
                        cell.font = Font(bold=True, size=10)
                        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    # Style data cells
                    elif r_idx > 1 and c_idx > 1:
                        # Extract course code and apply color
                        course_code = self._extract_course_code(value)
                        
                        if value == 'LUNCH BREAK':
                            cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                            cell.font = Font(bold=True, color='FFFFFF')
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        elif value == 'Free':
                            cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        elif course_code:
                            color = self._get_course_color(course_code)
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                            cell.font = Font(size=10)
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    # Apply border
                    cell.border = thin_border
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 12  # Day column
            for col in range(2, ws.max_column + 1):
                ws.column_dimensions[ws.cell(1, col).column_letter].width = 25
            
            # Adjust row heights
            ws.row_dimensions[1].height = 30  # Header row
            for row in range(2, ws.max_row + 1):
                ws.row_dimensions[row].height = 35
        
        # Save workbook
        output_path = os.path.join(self.input_dir, output_file)
        wb.save(output_path)
        
        print(f"\n✅ Excel file created: {output_path}")
        print(f"📊 Sheets: {len(wb.sheetnames)}")
        print(f"🎨 Unique courses colored: {len(self.course_colors)}")
        
        return output_path
    
    def export_single_timetable(self, csv_filename, output_filename=None):
        """Export a single timetable CSV to Excel"""
        csv_path = os.path.join(self.input_dir, csv_filename)
        
        if not os.path.exists(csv_path):
            print(f"CSV file not found: {csv_path}")
            return None
        
        if output_filename is None:
            output_filename = csv_filename.replace('.csv', '.xlsx')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = Path(csv_filename).stem.replace('_Timetable', '').replace('_', ' ')[:31]
        
        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Read CSV
        df = pd.read_csv(csv_path, index_col=0)
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Style header row
                if r_idx == 1:
                    cell.font = Font(bold=True, color='FFFFFF', size=11)
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                # Style index column (days)
                elif c_idx == 1 and r_idx > 1:
                    cell.font = Font(bold=True, size=10)
                    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                # Style data cells
                elif r_idx > 1 and c_idx > 1:
                    # Extract course code and apply color
                    course_code = self._extract_course_code(value)
                    
                    if value == 'LUNCH BREAK':
                        cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif value == 'Free':
                        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif course_code:
                        color = self._get_course_color(course_code)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        cell.font = Font(size=10)
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Apply border
                cell.border = thin_border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12  # Day column
        for col in range(2, ws.max_column + 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = 25
        
        # Adjust row heights
        ws.row_dimensions[1].height = 30  # Header row
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 35
        
        # Save workbook
        output_path = os.path.join(self.input_dir, output_filename)
        wb.save(output_path)
        
        print(f"✅ Excel file created: {output_path}")
        return output_path

if __name__ == '__main__':
    exporter = TimetableExcelExporter()
    exporter.export_to_excel()
